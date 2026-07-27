from __future__ import annotations

import argparse
import hashlib
import io
import json
import mimetypes
import os
import shutil
import sys
import tempfile
import threading
import time
import uuid
import webbrowser
import zipfile
from email.parser import BytesParser
from email.policy import default
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from .database import Database, utc_now, validate_state
from .connectors import ConnectorError
from .diagnostics import DiagnosticLogger, app_version
from .defaults import restore_default_accounts
from .integration import ConnectorService, ensure_connector_defaults
from .service import VoucherService
from .security import SecretStore
from .runtime import EnvironmentService, LauncherClient
from .setup import (
    CAPABILITY_CATALOG,
    SetupService,
    empty_production_state,
    ensure_state_v2,
    invalidate_if_upstream_changed,
)


MAX_REQUEST_BYTES = 200 * 1024 * 1024


class JobManager:
    def __init__(self, diagnostics: DiagnosticLogger | None = None) -> None:
        self.jobs: dict[str, dict] = {}
        self.threads: dict[str, threading.Thread] = {}
        self.lock = threading.Lock()
        self.diagnostics = diagnostics

    def submit(self, operation: str, target, *args, **kwargs) -> dict:
        job_id = f"JOB-{uuid.uuid4().hex[:12].upper()}"
        job = {
            "id": job_id,
            "operation": operation,
            "status": "queued",
            "progress": {
                "processed": 0,
                "total": len(args[0]) if args and isinstance(args[0], (list, tuple)) else 1,
                "percent": 0,
            },
            "createdAt": utc_now(),
        }
        with self.lock:
            self.jobs[job_id] = job
        if self.diagnostics:
            self.diagnostics.log(
                "INFO",
                "background_job",
                "JOB_QUEUED",
                f"后台任务已进入队列：{operation}",
                correlation_id=job_id,
                operation=operation,
                context={"total": job["progress"]["total"]},
            )

        def update(progress: dict) -> None:
            with self.lock:
                job["progress"] = progress
                job["updatedAt"] = utc_now()

        def run() -> None:
            with self.lock:
                job["status"] = "running"
                job["startedAt"] = utc_now()
            started = time.monotonic()
            if self.diagnostics:
                self.diagnostics.log(
                    "INFO",
                    "background_job",
                    "JOB_STARTED",
                    f"后台任务开始执行：{operation}",
                    correlation_id=job_id,
                    operation=operation,
                )
            try:
                result = target(*args, progress=update, **kwargs)
                with self.lock:
                    job["status"] = "completed"
                    job["result"] = result
                    job["finishedAt"] = utc_now()
                if self.diagnostics:
                    self.diagnostics.log(
                        "INFO",
                        "background_job",
                        "JOB_COMPLETED",
                        f"后台任务执行完成：{operation}",
                        correlation_id=job_id,
                        operation=operation,
                        duration_ms=round((time.monotonic() - started) * 1000),
                        context={"progress": job["progress"]},
                    )
            except Exception as exc:
                with self.lock:
                    job["status"] = "failed"
                    job["error"] = str(exc)
                    job["finishedAt"] = utc_now()
                if self.diagnostics:
                    self.diagnostics.exception(
                        "background_job",
                        "JOB_FAILED",
                        f"后台任务执行失败：{operation}",
                        exc,
                        user_action="打开诊断日志复制该任务编号，检查失败原因后重试",
                        correlation_id=job_id,
                        operation=operation,
                        duration_ms=round((time.monotonic() - started) * 1000),
                    )
            finally:
                with self.lock:
                    self.threads.pop(job_id, None)

        thread = threading.Thread(target=run, name=job_id, daemon=True)
        with self.lock:
            self.threads[job_id] = thread
            try:
                thread.start()
            except Exception:
                self.threads.pop(job_id, None)
                self.jobs.pop(job_id, None)
                raise
        return self.get(job_id)

    def get(self, job_id: str) -> dict:
        with self.lock:
            job = self.jobs.get(job_id)
            if not job:
                raise ValueError("后台任务不存在或已过期")
            return json.loads(json.dumps(job, ensure_ascii=False))

    def status(self) -> dict:
        with self.lock:
            active = [
                {
                    "id": item["id"],
                    "operation": item["operation"],
                    "status": item["status"],
                    "progress": item.get("progress", {}),
                }
                for item in self.jobs.values()
                if item["status"] in {"queued", "running"}
            ]
        return {"active": active, "activeCount": len(active)}

    def wait_for_idle(self, timeout: float = 5.0) -> bool:
        deadline = time.monotonic() + max(timeout, 0)
        while True:
            with self.lock:
                threads = list(self.threads.values())
            if not threads:
                return True
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                return False
            for thread in threads:
                thread.join(min(remaining, 0.1))


def make_handler(
    database: Database,
    static_dir: Path,
    diagnostics: DiagnosticLogger | None = None,
    host: str = "127.0.0.1",
    port: int = 8765,
):
    diagnostics = diagnostics or DiagnosticLogger(database)
    service = VoucherService(database)
    secret_store = SecretStore()
    connector_service = ConnectorService(database, secret_store)
    setup_service = SetupService(database)
    jobs = JobManager(diagnostics)
    environment_service = EnvironmentService(database, secret_store, static_dir, host, port)
    launcher_client = LauncherClient(environment_service.store)

    def runtime_status() -> dict:
        job_status = jobs.status()
        state = database.get_state() or {}
        active_outbox = [
            item for item in state.get("outbox", [])
            if item.get("status") in {"processing", "sending"}
        ]
        blockers = []
        if job_status["activeCount"]:
            blockers.append(f"有 {job_status['activeCount']} 个后台任务正在运行")
        if active_outbox:
            blockers.append(f"有 {len(active_outbox)} 个凭证推送正在处理")
        return {
            "serviceHealthy": True,
            "coreVersion": app_version(),
            "launcherVersion": os.environ.get("AUTO_VOUCHER_LAUNCHER_VERSION", ""),
            "databaseStatus": database.quick_check(),
            "databaseVersion": database.schema_version(),
            "stateVersion": int(state.get("version") or 2),
            "host": host,
            "port": port,
            "jobs": job_status,
            "restartAllowed": not blockers,
            "restartBlockers": blockers,
        }

    def initialize_production_state() -> dict:
        state = database.get_state()
        if state is None:
            state = empty_production_state()
            return database.put_state(state)
        if int(state.get("version") or 0) < 2:
            backup_dir = database.data_dir / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup = create_backup_package(database, state)
            backup_path = backup_dir / f"pre-v2-migration-{uuid.uuid4().hex[:12]}.zip"
            backup_path.write_bytes(backup)
            database.clear_business_data()
            for child in list(database.archive_dir.iterdir()):
                shutil.rmtree(child) if child.is_dir() else child.unlink()
            for connector_id, names in {
                "feishu-approval": ("app_secret",),
                "kingdee-k3cloud": ("password",),
                "yonyou-u8": ("access_token",),
                "inspur-gscloud": ("access_token",),
            }.items():
                for name in names:
                    try:
                        secret_store.delete(connector_id, name)
                    except RuntimeError:
                        pass
            return database.put_state(empty_production_state())
        if ensure_state_v2(state):
            state = database.put_state(state)
        return state

    class Handler(BaseHTTPRequestHandler):
        server_version = "AutoVoucher/0.2"
        job_manager = jobs

        def handle_one_request(self) -> None:
            self._request_id = f"REQ-{uuid.uuid4().hex[:12].upper()}"
            self._request_started = time.monotonic()
            self._response_started = False
            try:
                super().handle_one_request()
            except Exception as exc:
                diagnostics.exception(
                    "http",
                    "UNHANDLED_REQUEST_ERROR",
                    "本地接口发生未处理异常",
                    exc,
                    user_action="复制支持编号并导出最近 7 天诊断包",
                    correlation_id=self._request_id,
                    operation=getattr(self, "command", ""),
                    context={"path": getattr(self, "path", "")},
                    duration_ms=round((time.monotonic() - self._request_started) * 1000),
                )
                if self._response_started:
                    self.close_connection = True
                else:
                    try:
                        self.json_response(
                            {"error": "本地服务发生未处理异常，请在诊断日志中复制支持编号"},
                            HTTPStatus.INTERNAL_SERVER_ERROR,
                        )
                    except Exception:
                        pass

        def send_response(self, code: int, message: str | None = None) -> None:
            self._response_started = True
            super().send_response(code, message)

        def do_GET(self) -> None:
            parsed_request = urlparse(self.path)
            path = parsed_request.path
            query = parse_qs(parsed_request.query)
            if path == "/api/health":
                state = database.get_state() or {}
                self.json_response({
                    "ok": True,
                    "storage": "sqlite",
                    "dataDir": str(database.data_dir),
                    "coreVersion": app_version(),
                    "databaseStatus": database.quick_check(),
                    "databaseVersion": database.schema_version(),
                    "stateVersion": int(state.get("version") or 2),
                    "staticAssets": (static_dir / "index.html").is_file(),
                })
            elif path == "/api/runtime/status":
                self.json_response(runtime_status())
            elif path == "/api/environment/status":
                self.json_response(environment_service.status())
            elif path == "/api/update/status":
                self.json_response(launcher_client.status())
            elif path == "/api/diagnostics/copy-summary":
                self.json_response(diagnostics.copy_summary(
                    environment=environment_service.status(),
                    update=launcher_client.status(),
                    runtime=runtime_status(),
                ))
            elif path == "/api/diagnostics/logs":
                self.json_response(diagnostics.query(
                    level=(query.get("level") or [""])[0],
                    category=(query.get("category") or [""])[0],
                    search=(query.get("search") or [""])[0],
                    correlation_id=(query.get("correlationId") or [""])[0],
                    date_from=(query.get("from") or [""])[0],
                    date_to=(query.get("to") or [""])[0],
                    limit=int((query.get("limit") or ["200"])[0]),
                    offset=int((query.get("offset") or ["0"])[0]),
                ))
            elif path == "/api/diagnostics/summary":
                self.json_response(diagnostics.summary(
                    int((query.get("hours") or ["24"])[0])
                ))
            elif path == "/api/diagnostics/export":
                self.send_diagnostics(
                    int((query.get("days") or ["7"])[0])
                )
            elif path == "/api/state":
                self.json_response({"state": initialize_production_state()})
            elif path == "/api/setup/catalog":
                self.json_response({"catalog": CAPABILITY_CATALOG})
            elif path.startswith("/api/jobs/"):
                try:
                    self.json_response({"job": jobs.get(path.removeprefix("/api/jobs/"))})
                except ValueError as exc:
                    self.json_response({"error": str(exc)}, HTTPStatus.NOT_FOUND)
            elif path == "/api/security/keyring":
                try:
                    self.json_response(secret_store.status())
                except RuntimeError as exc:
                    self.json_response({"available": False, "error": str(exc)}, HTTPStatus.SERVICE_UNAVAILABLE)
            elif path == "/api/backup":
                self.send_backup()
            elif path.startswith("/api/vouchers/") and path.endswith("/export.xlsx"):
                voucher_id = path.removeprefix("/api/vouchers/").removesuffix("/export.xlsx")
                self.send_voucher_xlsx(voucher_id)
            elif path.startswith("/api/"):
                self.json_response({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)
            else:
                self.send_static(path)

        def do_PUT(self) -> None:
            if not self.allow_local_request():
                return
            path = urlparse(self.path).path
            try:
                if path == "/api/state":
                    state = self.read_json().get("state")
                    if isinstance(state, dict):
                        ensure_state_v2(state)
                        invalidate_if_upstream_changed(database.get_state(), state)
                    self.json_response({"state": database.put_state(state)})
                elif path == "/api/diagnostics/settings":
                    payload = self.read_json()
                    self.json_response({"settings": diagnostics.update_settings(
                        int(payload.get("retentionDays") or 30),
                        int(payload.get("maxEntries") or 50_000),
                    )})
                elif path.startswith("/api/connectors/") and path.endswith("/config"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/config").strip("/")
                    payload = self.read_json()
                    self.json_response(connector_service.configure(
                        connector_id,
                        payload.get("config") or {},
                        str(payload.get("productionConfirmation") or ""),
                    ))
                else:
                    self.json_response({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)
            except (ValueError, json.JSONDecodeError) as exc:
                self.json_response({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        def do_POST(self) -> None:
            if not self.allow_local_request():
                return
            path = urlparse(self.path).path
            try:
                if path == "/api/diagnostics/client":
                    payload = self.read_json()
                    entry = diagnostics.log(
                        str(payload.get("level") or "ERROR"),
                        "frontend",
                        str(payload.get("eventCode") or "FRONTEND_EVENT"),
                        str(payload.get("message") or "浏览器端事件"),
                        user_action=str(payload.get("userAction") or "刷新页面；如问题持续，请导出诊断包"),
                        correlation_id=str(payload.get("correlationId") or self._request_id),
                        operation=str(payload.get("operation") or ""),
                        context=payload.get("context") or {},
                        error=payload.get("error") or {},
                        source="frontend",
                    )
                    self.json_response({"ok": True, "id": entry["id"]}, HTTPStatus.CREATED)
                elif path == "/api/import":
                    files, fields = self.read_multipart()
                    selected = files.get("files", [])
                    if not selected:
                        raise ValueError("没有收到文件")
                    mapping = json.loads(fields.get("mapping", "{}"))
                    if not isinstance(mapping, dict):
                        raise ValueError("字段映射格式无效")
                    job = jobs.submit(
                        "import",
                        service.import_files,
                        selected,
                        mapping=mapping or None,
                        template_name=fields.get("templateName", "").strip(),
                    )
                    self.json_response({"job": job}, HTTPStatus.ACCEPTED)
                elif path == "/api/import/preview":
                    files, _ = self.read_multipart()
                    selected = files.get("file", [])
                    if len(selected) != 1:
                        raise ValueError("请选择一个文件预览")
                    filename, content, _media_type = selected[0]
                    self.json_response(service.preview_file(filename, content))
                elif path == "/api/master-data/accounts/restore-defaults":
                    self.read_json()
                    state = initialize_production_state()
                    count = restore_default_accounts(state, utc_now())
                    state.setdefault("auditLog", []).insert(0, {
                        "id": f"LOG-{uuid.uuid4().hex[:10].upper()}",
                        "action": "恢复默认科目",
                        "subject": "小企业会计准则科目表",
                        "operator": state.get("operator") or "本机操作者",
                        "detail": f"停用当前科目版本并恢复 {count} 个内置默认科目",
                        "at": utc_now(),
                    })
                    self.json_response({"state": database.put_state(state), "count": count})
                elif path == "/api/setup/plan":
                    self.json_response(setup_service.plan(self.read_json()))
                elif path == "/api/setup/preflight":
                    self.read_json()
                    self.json_response(setup_service.preflight())
                elif path == "/api/setup/activate":
                    environment = environment_service.status()
                    environment_service.assert_production_ready()
                    payload = self.read_json()
                    payload["environmentValidation"] = {
                        "supportCode": environment.get("supportCode"),
                        "checkedAt": environment.get("checkedAt"),
                        "coreVersion": environment.get("coreVersion"),
                        "overallStatus": environment.get("overallStatus"),
                        "productionFingerprint": environment.get("productionFingerprint"),
                    }
                    self.json_response(setup_service.activate(payload))
                elif path == "/api/environment/check":
                    payload = self.read_json()
                    self.json_response(environment_service.check(
                        include_network=payload.get("includeNetwork", True) is not False,
                        browser_checks=payload.get("browserChecks")
                        if isinstance(payload.get("browserChecks"), list)
                        else None,
                    ))
                elif path == "/api/environment/repair":
                    payload = self.read_json()
                    action = str(payload.get("action") or "")
                    if action == "recreate-shortcut":
                        self.json_response(launcher_client.command(action))
                    else:
                        self.json_response(environment_service.repair(action))
                elif path == "/api/update/check":
                    self.read_json()
                    self.json_response(launcher_client.command("check"))
                elif path == "/api/update/download":
                    self.read_json()
                    self.json_response(launcher_client.command("download"))
                elif path == "/api/update/apply":
                    self.read_json()
                    runtime = runtime_status()
                    if not runtime["restartAllowed"]:
                        raise ValueError("当前不能重启更新：" + "；".join(runtime["restartBlockers"]))
                    backup_dir = database.data_dir / "backups"
                    backup_dir.mkdir(parents=True, exist_ok=True)
                    backup_path = backup_dir / f"pre-update-{uuid.uuid4().hex[:12]}.sqlite3"
                    backup_content = database.backup_bytes()
                    backup_path.write_bytes(backup_content)
                    self.json_response(launcher_client.command("apply", {
                        "databaseBackup": str(backup_path),
                        "databaseSha256": hashlib.sha256(backup_content).hexdigest(),
                    }))
                elif path == "/api/update/postpone":
                    self.json_response(launcher_client.command("postpone", self.read_json()))
                elif path == "/api/setup/reset":
                    payload = self.read_json()
                    if str(payload.get("confirmation") or "") != "备份并全量初始化":
                        raise ValueError("全量初始化必须明确输入“备份并全量初始化”")
                    current = initialize_production_state()
                    backup_dir = database.data_dir / "backups"
                    backup_dir.mkdir(parents=True, exist_ok=True)
                    backup = create_backup_package(database, current)
                    backup_path = backup_dir / f"manual-reset-{uuid.uuid4().hex[:12]}.zip"
                    backup_path.write_bytes(backup)
                    database.clear_business_data()
                    for child in list(database.archive_dir.iterdir()):
                        shutil.rmtree(child) if child.is_dir() else child.unlink()
                    cleared_secrets = []
                    for connector in current.get("connectors", []):
                        names = ("app_secret",) if connector.get("adapter") == "feishu-approval-v4" else ("password", "access_token")
                        for name in names:
                            try:
                                secret_store.delete(str(connector.get("id") or ""), name)
                                cleared_secrets.append(f"{connector.get('id')}:{name}")
                            except RuntimeError:
                                continue
                    fresh = database.put_state(empty_production_state())
                    self.json_response({
                        "state": fresh,
                        "backup": {
                            "path": str(backup_path),
                            "sha256": hashlib.sha256(backup).hexdigest(),
                            "bytes": len(backup),
                        },
                        "clearedSecrets": cleared_secrets,
                    })
                elif path == "/api/templates/preview":
                    files, fields = self.read_multipart()
                    selected = files.get("file", [])
                    if len(selected) != 1:
                        raise ValueError("请选择一个 ERP 空白模板或成功样例")
                    filename, content, _media_type = selected[0]
                    self.json_response({"preview": setup_service.preview_template(
                        filename,
                        content,
                        fields.get("targetSystemId", "").strip(),
                    )})
                elif path == "/api/templates/validate":
                    self.json_response(setup_service.validate_template(self.read_json()))
                elif path.startswith("/api/vouchers/") and path.endswith("/push"):
                    voucher_id = path.removeprefix("/api/vouchers/").removesuffix("/push")
                    payload = self.read_json()
                    connector_id = str(payload.get("connectorId") or "")
                    if not connector_id:
                        raise ValueError("请选择已验证的目标 ERP 连接器")
                    if str(payload.get("expectedEnvironment") or "") == "生产环境":
                        environment_service.assert_production_ready()
                    self.json_response(connector_service.push_voucher(
                        voucher_id,
                        connector_id,
                        str(payload.get("expectedEnvironment") or ""),
                    ))
                elif path.startswith("/api/vouchers/") and path.endswith("/recheck"):
                    voucher_id = path.removeprefix("/api/vouchers/").removesuffix("/recheck")
                    payload = self.read_json()
                    connector_id = str(payload.get("connectorId") or "")
                    if not connector_id:
                        raise ValueError("请选择目标 ERP 连接器")
                    self.json_response(connector_service.recheck_voucher(voucher_id, connector_id))
                elif path.startswith("/api/vouchers/") and path.endswith("/preflight"):
                    voucher_id = path.removeprefix("/api/vouchers/").removesuffix("/preflight")
                    payload = self.read_json()
                    self.json_response(connector_service.preflight(
                        voucher_id,
                        str(payload.get("connectorId") or ""),
                        str(payload.get("expectedEnvironment") or ""),
                    ))
                elif path.startswith("/api/connectors/") and path.endswith("/test"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/test").strip("/")
                    self.read_json()
                    self.json_response(connector_service.probe(connector_id))
                elif path.startswith("/api/connectors/") and path.endswith("/sync-approvals"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/sync-approvals").strip("/")
                    self.read_json()
                    job = jobs.submit(
                        "sync-approvals",
                        lambda target_id, progress: connector_service.sync_approvals(target_id),
                        connector_id,
                    )
                    self.json_response({"job": job}, HTTPStatus.ACCEPTED)
                elif path.startswith("/api/connectors/") and path.endswith("/sync-master-data"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/sync-master-data").strip("/")
                    self.read_json()
                    job = jobs.submit(
                        "sync-master-data",
                        lambda target_id, progress: connector_service.sync_master_data(target_id),
                        connector_id,
                    )
                    self.json_response({"job": job}, HTTPStatus.ACCEPTED)
                elif path.startswith("/api/connectors/") and path.endswith("/query-voucher"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/query-voucher").strip("/")
                    payload = self.read_json()
                    self.json_response(connector_service.query_external_voucher(
                        connector_id,
                        number=str(payload.get("number") or ""),
                        reference=str(payload.get("reference") or ""),
                    ))
                elif path.startswith("/api/connectors/") and path.endswith("/query-ledger"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/query-ledger").strip("/")
                    payload = self.read_json()
                    self.json_response(connector_service.query_external_ledger(
                        connector_id,
                        {
                            "ledger": str(payload.get("ledger") or ""),
                            "period": str(payload.get("period") or ""),
                            "account": str(payload.get("account") or ""),
                            "dimension": str(payload.get("dimension") or ""),
                        },
                    ))
                elif path.startswith("/api/connectors/") and path.endswith("/query-report"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/query-report").strip("/")
                    payload = self.read_json()
                    self.json_response(connector_service.query_external_report(
                        connector_id,
                        str(payload.get("reportType") or ""),
                        str(payload.get("period") or ""),
                    ))
                elif path.startswith("/api/connectors/") and path.endswith("/secret"):
                    connector_id = path.removeprefix("/api/connectors/").removesuffix("/secret").strip("/")
                    payload = self.read_json()
                    secret_store.set(
                        connector_id,
                        str(payload.get("name", "")).strip(),
                        str(payload.get("value", "")),
                    )
                    self.json_response({"ok": True, "storedIn": "os-keyring"})
                elif path == "/api/restore":
                    files, _ = self.read_multipart()
                    uploaded = files.get("backup", [])
                    if len(uploaded) != 1:
                        raise ValueError("请选择一个有效备份包")
                    state = restore_backup(database, uploaded[0][1])
                    self.json_response({"state": state})
                elif path == "/api/restore/preview":
                    files, _ = self.read_multipart()
                    uploaded = files.get("backup", [])
                    if len(uploaded) != 1:
                        raise ValueError("请选择一个有效备份包")
                    self.json_response({"report": inspect_backup(uploaded[0][1])})
                else:
                    self.json_response({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)
            except ConnectorError as exc:
                action_by_category = {
                    "authentication": "重新保存密钥并测试连接",
                    "permission": "联系系统管理员检查专用集成账号权限",
                    "rate_limit": "等待片刻后重试；不要连续点击",
                    "network": "检查目标地址、网络和证书后重试",
                    "master_data": "同步目标系统基础资料并核对编码",
                    "validation": "按错误详情修正字段映射或业务数据",
                    "security": "改用专用最小权限账号",
                }
                diagnostics.log(
                    "ERROR" if not exc.retryable else "WARNING",
                    "connector",
                    f"CONNECTOR_{exc.code}",
                    exc.message,
                    user_action=action_by_category.get(
                        exc.category,
                        "根据提示修正后重新测试；如问题持续，请导出诊断包",
                    ),
                    correlation_id=getattr(self, "_request_id", ""),
                    operation=path,
                    context={"connectorError": exc.as_dict()},
                )
                self.json_response(
                    {"error": exc.message, "connectorError": exc.as_dict()},
                    HTTPStatus.BAD_GATEWAY if exc.retryable else HTTPStatus.BAD_REQUEST,
                )
            except (ValueError, zipfile.BadZipFile, json.JSONDecodeError) as exc:
                self.json_response({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        def do_DELETE(self) -> None:
            if not self.allow_local_request():
                return
            if urlparse(self.path).path == "/api/state":
                self.json_response(
                    {"error": "不允许直接删除工作区；请使用“备份并全量初始化”完成受控重置"},
                    HTTPStatus.METHOD_NOT_ALLOWED,
                )
            else:
                self.json_response({"error": "接口不存在"}, HTTPStatus.NOT_FOUND)

        def read_body(self) -> bytes:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0 or length > MAX_REQUEST_BYTES:
                raise ValueError("请求为空或超过 200 MB")
            return self.rfile.read(length)

        def allow_local_request(self) -> bool:
            host = self.headers.get("Host", "").split(":", 1)[0].strip("[]")
            if host not in {"127.0.0.1", "localhost"}:
                self.json_response({"error": "拒绝非本机 Host"}, HTTPStatus.FORBIDDEN)
                return False
            origin = self.headers.get("Origin")
            if origin:
                parsed = urlparse(origin)
                if parsed.hostname not in {"127.0.0.1", "localhost"}:
                    self.json_response({"error": "拒绝跨站写入"}, HTTPStatus.FORBIDDEN)
                    return False
            return True

        def read_json(self) -> dict:
            return json.loads(self.read_body().decode("utf-8"))

        def read_multipart(self) -> tuple[dict[str, list[tuple[str, bytes, str]]], dict[str, str]]:
            content_type = self.headers.get("Content-Type", "")
            if not content_type.startswith("multipart/form-data"):
                raise ValueError("请求必须使用 multipart/form-data")
            message = BytesParser(policy=default).parsebytes(
                b"Content-Type: " + content_type.encode("ascii") + b"\r\nMIME-Version: 1.0\r\n\r\n" + self.read_body()
            )
            files: dict[str, list[tuple[str, bytes, str]]] = {}
            fields: dict[str, str] = {}
            for part in message.iter_parts():
                name = part.get_param("name", header="content-disposition")
                if not name:
                    continue
                filename = part.get_filename()
                if filename:
                    safe_name = Path(filename).name
                    files.setdefault(name, []).append(
                        (safe_name, part.get_payload(decode=True), part.get_content_type())
                    )
                else:
                    charset = part.get_content_charset() or "utf-8"
                    fields[name] = part.get_payload(decode=True).decode(charset, errors="replace").strip()
            return files, fields

        def json_response(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
            payload = dict(payload)
            if int(status) >= 400:
                payload.setdefault("correlationId", getattr(self, "_request_id", ""))
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("X-Correlation-ID", getattr(self, "_request_id", ""))
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
            self._record_response(status, payload)

        def _record_response(self, status: HTTPStatus, payload: dict) -> None:
            path = urlparse(getattr(self, "path", "")).path
            if (
                int(status) < 400
                and (
                    path in {"/api/state", "/api/health", "/api/diagnostics/logs", "/api/diagnostics/summary"}
                    or path.startswith("/api/jobs/")
                    or path == "/api/diagnostics/client"
                )
            ):
                return
            level = "ERROR" if int(status) >= 500 else "WARNING" if int(status) >= 400 else "INFO"
            category = (
                "connector" if "/connectors/" in path
                else "import" if "/import" in path
                else "backup" if path.startswith("/api/backup") or path.startswith("/api/restore")
                else "security" if "/security/" in path
                else "http"
            )
            diagnostics.log(
                level,
                category,
                f"HTTP_{getattr(self, 'command', 'REQUEST')}_{int(status)}",
                f"{getattr(self, 'command', 'REQUEST')} {path} 返回 {int(status)}",
                user_action=(
                    "根据错误提示修正后重试；如问题持续，请复制支持编号"
                    if int(status) >= 400 else ""
                ),
                correlation_id=getattr(self, "_request_id", ""),
                operation=getattr(self, "command", ""),
                context={
                    "path": path,
                    "status": int(status),
                    "error": payload.get("error") if int(status) >= 400 else "",
                    "connectorError": payload.get("connectorError") if int(status) >= 400 else {},
                },
                duration_ms=round(
                    (time.monotonic() - getattr(self, "_request_started", time.monotonic())) * 1000
                ),
            )

        def send_backup(self) -> None:
            state = database.get_state()
            if state is None:
                return self.json_response({"error": "没有可备份的数据"}, HTTPStatus.NOT_FOUND)
            body = create_backup_package(database, state)
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", 'attachment; filename="AutoVoucher-backup.zip"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            diagnostics.log(
                "INFO",
                "backup",
                "BACKUP_EXPORTED",
                "本地完整备份已生成",
                correlation_id=getattr(self, "_request_id", ""),
                context={"bytes": len(body)},
            )

        def send_diagnostics(self, days: int) -> None:
            body, support_code = diagnostics.export_zip(database.get_state(), days, {
                "environment": environment_service.status(),
                "update": launcher_client.status(),
                "runtime": runtime_status(),
                "launcher": launcher_client.diagnostics(),
            })
            filename = f"AutoVoucher-diagnostics-{support_code}.zip"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("X-Support-Code", support_code)
            self.send_header("X-Correlation-ID", getattr(self, "_request_id", ""))
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def send_voucher_xlsx(self, voucher_id: str) -> None:
            state = database.get_state() or {}
            if not (state.get("productionActivation") or {}).get("enabled"):
                return self.json_response(
                    {"error": "尚未通过测试上线，禁止生产导出凭证"},
                    HTTPStatus.CONFLICT,
                )
            voucher = next((item for item in state.get("vouchers", []) if item.get("id") == voucher_id), None)
            if not voucher:
                return self.json_response({"error": "凭证不存在"}, HTTPStatus.NOT_FOUND)
            if voucher.get("status") not in {"已确认", "已推送"}:
                return self.json_response({"error": "只有已确认或已推送凭证可以导出"}, HTTPStatus.CONFLICT)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                return self.json_response({"error": "缺少 XLSX 导出组件 openpyxl"}, HTTPStatus.SERVICE_UNAVAILABLE)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "凭证"
            headers = ["凭证号", "凭证日期", "摘要", "科目编码", "科目名称", "借方", "贷方", "部门", "项目", "供应商"]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="FF5A5F")
            for line in voucher.get("lines", []):
                dimensions = line.get("dimensions") or {}
                sheet.append([
                    voucher.get("number"),
                    voucher.get("accountingDate"),
                    line.get("summary"),
                    line.get("accountCode"),
                    line.get("accountName"),
                    (line.get("debitCents") or 0) / 100,
                    (line.get("creditCents") or 0) / 100,
                    dimensions.get("department") or "",
                    dimensions.get("project") or "",
                    dimensions.get("supplier") or "",
                ])
            for column in ("F", "G"):
                for cell in sheet[column][1:]:
                    cell.number_format = '#,##0.00'
            widths = [18, 14, 30, 12, 18, 14, 14, 16, 18, 24]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[chr(64 + index)].width = width
            output = io.BytesIO()
            workbook.save(output)
            body = output.getvalue()
            digest = hashlib.sha256(body).hexdigest()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="voucher.xlsx"')
            self.send_header("X-Content-SHA256", digest)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def send_static(self, request_path: str) -> None:
            relative = unquote(request_path).lstrip("/") or "index.html"
            target = (static_dir / relative).resolve()
            if static_dir.resolve() not in target.parents and target != static_dir.resolve():
                return self.send_error(HTTPStatus.FORBIDDEN)
            if not target.is_file():
                target = static_dir / "index.html"
            if not target.is_file():
                return self.json_response(
                    {"error": "前端尚未构建，请先运行 npm run build"},
                    HTTPStatus.SERVICE_UNAVAILABLE,
                )
            body = target.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "application/octet-stream")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, fmt: str, *args) -> None:
            return

    return Handler


def restore_backup(database: Database, content: bytes) -> dict:
    manifest, state, members = verify_backup(content)
    current_state = database.get_state()
    if current_state:
        current_audit = current_state.get("auditLog", [])
        current_ids = {item.get("id") for item in current_audit}
        restored_only_audit = [
            item for item in state.get("auditLog", [])
            if item.get("id") not in current_ids
        ]
        state["auditLog"] = [
            {
                "id": f"LOG-RESTORE-{uuid.uuid4().hex[:10].upper()}",
                "action": "恢复备份",
                "subject": "本地工作空间",
                "operator": current_state.get("operator") or "本机操作者",
                "detail": (
                    f"完整性校验通过；恢复 {len(state.get('sourceDocuments', []))} 份原始资料，"
                    f"保留 {len(current_audit)} 条既有审计记录"
                ),
                "at": utc_now(),
            },
            *current_audit,
            *restored_only_audit,
        ]
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        with tempfile.TemporaryDirectory(prefix="auto-voucher-restore-") as temporary:
            root = Path(temporary)
            for name in members:
                if not name.startswith("archive/") or name.endswith("/"):
                    continue
                target = (root / name).resolve()
                if root.resolve() not in target.parents:
                    raise ValueError("备份包含不安全路径")
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(archive.read(name))
            restored_archive = root / "archive"
            if restored_archive.exists():
                for source in restored_archive.rglob("*"):
                    if source.is_file():
                        destination = database.archive_dir / source.relative_to(restored_archive)
                        destination.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(source, destination)
        restored_state = database.put_state(state)
        for document in restored_state.get("sourceDocuments", []):
            digest = document.get("fullHash")
            if not digest or database.has_source(digest):
                continue
            candidates = list((database.archive_dir / digest[:2]).glob(f"{digest}.*"))
            if candidates:
                database.register_source(
                    digest,
                    document.get("name") or candidates[0].name,
                    "application/octet-stream",
                    int(document.get("size") or candidates[0].stat().st_size),
                    candidates[0],
                )
        return restored_state


def create_backup_package(database: Database, state: dict) -> bytes:
    state_bytes = json.dumps(state, ensure_ascii=False, indent=2).encode("utf-8")
    members: dict[str, bytes] = {
        "state.json": state_bytes,
        "database/auto-voucher.sqlite3": database.backup_bytes(),
    }
    for source in database.archive_dir.rglob("*"):
        if source.is_file():
            members[source.relative_to(database.data_dir).as_posix()] = source.read_bytes()
    manifest = {
        "kind": "auto-voucher-backup",
        "version": 2,
        "createdAt": utc_now(),
        "scope": {
            "sourceDocuments": len(state.get("sourceDocuments", [])),
            "events": len(state.get("events", [])),
            "vouchers": len(state.get("vouchers", [])),
            "rules": len(state.get("rules", [])),
            "auditEvents": len(state.get("auditLog", [])),
        },
        "files": {
            name: {"sha256": hashlib.sha256(payload).hexdigest(), "size": len(payload)}
            for name, payload in members.items()
        },
    }
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        for name, payload in members.items():
            archive.writestr(name, payload)
    return output.getvalue()


def verify_backup(content: bytes) -> tuple[dict, dict, list[str]]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        try:
            manifest = json.loads(archive.read("manifest.json"))
        except KeyError as exc:
            raise ValueError("备份缺少 manifest.json") from exc
        if manifest.get("kind") != "auto-voucher-backup":
            raise ValueError("不是有效的 Auto Voucher 备份包")
        if manifest.get("version") != 2 or not isinstance(manifest.get("files"), dict):
            raise ValueError("备份版本过旧或缺少完整性清单，请使用当前版本重新备份")
        names = set(archive.namelist())
        for name, expected in manifest["files"].items():
            if name not in names:
                raise ValueError(f"备份缺少文件：{name}")
            payload = archive.read(name)
            digest = hashlib.sha256(payload).hexdigest()
            if digest != expected.get("sha256") or len(payload) != expected.get("size"):
                raise ValueError(f"备份文件完整性校验失败：{name}")
        state = json.loads(archive.read("state.json"))
        validate_state(state)
        return manifest, state, list(manifest["files"])


def inspect_backup(content: bytes) -> dict:
    manifest, state, _members = verify_backup(content)
    return {
        "valid": True,
        "version": manifest["version"],
        "createdAt": manifest.get("createdAt"),
        "scope": manifest.get("scope", {}),
        "company": state.get("company"),
        "ledger": state.get("ledger"),
        "message": "备份完整性校验通过，可进入恢复确认。",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Auto Voucher 本地凭证工作台")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=int(os.environ.get("AUTO_VOUCHER_PORT", "8765")))
    parser.add_argument("--data-dir", type=Path)
    parser.add_argument("--no-browser", action="store_true")
    args = parser.parse_args()
    if args.host not in {"127.0.0.1", "localhost"}:
        parser.error("默认版本只允许监听本机回环地址")
    database = Database(args.data_dir)
    diagnostics = DiagnosticLogger(database)
    bundled_root = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parents[2]))
    static_dir = bundled_root / "dist"
    handler = make_handler(database, static_dir, diagnostics, args.host, args.port)
    server = ThreadingHTTPServer(
        (args.host, args.port),
        handler,
    )
    url = f"http://{args.host}:{args.port}"
    print(f"Auto Voucher 已启动：{url}")
    print(f"数据目录：{database.data_dir}")
    diagnostics.log(
        "INFO",
        "application",
        "APPLICATION_STARTED",
        "Auto Voucher 本地服务已启动",
        context={"host": args.host, "port": args.port, "frozen": bool(getattr(sys, "frozen", False))},
    )
    if not args.no_browser:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        jobs_finished = handler.job_manager.wait_for_idle(5)
        if not jobs_finished:
            diagnostics.log(
                "WARNING",
                "background_job",
                "JOB_SHUTDOWN_TIMEOUT",
                "本地服务停止前仍有后台任务未在 5 秒内完成",
                user_action="重新启动后检查诊断日志和最近导入结果，再决定是否重试。",
                context=handler.job_manager.status(),
            )
        diagnostics.log(
            "INFO",
            "application",
            "APPLICATION_STOPPED",
            "Auto Voucher 本地服务已停止",
        )
        server.server_close()


if __name__ == "__main__":
    main()
