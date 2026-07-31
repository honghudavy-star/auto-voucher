from __future__ import annotations

import hashlib
import csv
import io
import json
import uuid
from pathlib import Path
from typing import Any

from .defaults import (
    DEFAULT_ACCOUNT_SOURCE,
    default_account_master_data,
    initialize_default_accounts,
)

from .database import Database, utc_now
from .importers import header_fingerprint, headers_for, parse_rows


STATE_VERSION = 2

APPROVAL_PROFILE_DEFAULTS = {
    "approvalName": "",
    "approvalFields": [],
    "fieldMapping": {},
    "fieldSources": [],
    "additionalApprovalFieldIds": [],
    "syncCursor": {},
}


def approval_profile_id(approval_code: str) -> str:
    digest = hashlib.sha256(str(approval_code or "").encode("utf-8")).hexdigest()[:12]
    return f"approval-{digest}"


def legacy_approval_profile(connector: dict[str, Any]) -> dict[str, Any] | None:
    approval_code = str(connector.get("approvalCode") or "").strip()
    if not approval_code:
        return None
    return {
        "id": approval_profile_id(approval_code),
        "approvalCode": approval_code,
        **{
            key: json.loads(json.dumps(connector.get(key, default), ensure_ascii=False))
            for key, default in APPROVAL_PROFILE_DEFAULTS.items()
        },
    }

KINGDEE_MASTER_DATA_QUERIES: list[dict[str, Any]] = [
    {"category": "organization", "categoryLabel": "组织", "formId": "ORG_Organizations", "fields": ["FNumber", "FName"]},
    {"category": "accountBook", "categoryLabel": "账簿", "formId": "BD_AccountBook", "fields": ["FNumber", "FName"]},
    {"category": "account", "categoryLabel": "科目", "formId": "BD_Account", "fields": ["FNumber", "FName"]},
    {"category": "customer", "categoryLabel": "客户", "formId": "BD_Customer", "fields": ["FNumber", "FName"]},
    {"category": "supplier", "categoryLabel": "供应商", "formId": "BD_Supplier", "fields": ["FNumber", "FName"]},
    {"category": "department", "categoryLabel": "部门", "formId": "BD_Department", "fields": ["FNumber", "FName"]},
    {"category": "employee", "categoryLabel": "员工", "formId": "BD_Empinfo", "fields": ["FNumber", "FName"]},
    {"category": "project", "categoryLabel": "项目", "formId": "BD_Project", "fields": ["FNumber", "FName"], "optional": True},
    {"category": "otherCounterparty", "categoryLabel": "其他往来", "formId": "FIN_OTHERS", "fields": ["FNumber", "FName"]},
    {"category": "assistantCategory", "categoryLabel": "辅助资料类别", "formId": "BOS_ASSISTANTDATA", "fields": ["FNumber", "FName"]},
    {
        "category": "assistantData",
        "categoryLabel": "辅助资料",
        "formId": "BOS_ASSISTANTDATA_DETAIL",
        "fields": ["FId", "FNumber", "FDataValue"],
        "idFields": ["FId", "FNumber"],
        "codeField": "FNumber",
        "nameField": "FDataValue",
        "replaceLegacyIdentity": True,
    },
    {"category": "dimensionDefinition", "categoryLabel": "核算维度定义", "formId": "BAS_FLEX", "fields": ["FNumber", "FName"]},
    {"category": "accountDimension", "categoryLabel": "科目核算维度", "formId": "BD_FLEXITEMPROPERTY", "fields": ["FNumber", "FName"]},
    {"category": "dimensionGroup", "categoryLabel": "核算维度组", "formId": "BD_FLEXITEMGROUP", "fields": ["FNumber", "FName"]},
    {
        "category": "dimensionValue",
        "categoryLabel": "核算维度值",
        "formId": "BD_FLEXITEMDETAILV",
        "fields": [
            "FFlex4.FNumber", "FFlex4.FName",
            "FFlex5.FNumber", "FFlex5.FName",
            "FFlex6.FNumber", "FFlex6.FName",
            "FFlex7.FNumber", "FFlex7.FName",
            "FFlex8.FNumber", "FFlex8.FName",
            "FFLEX9.FNumber", "FFLEX9.FName",
            "FFLEX11.FNumber", "FFLEX11.FName",
            "FFLEX14.FNumber", "FFLEX14.FName",
            "FFLEX15.FNumber", "FFLEX15.FName",
            "FFLEX16.FNumber", "FFLEX16.FName",
            "FF100002.FNumber", "FF100002.FName",
            "FF100003", "FF100004", "FF100006", "FF100007",
        ],
        "dimensionMappings": [
            {"category": "dimensionSupplier", "categoryLabel": "核算维度·供应商", "codeField": "FFlex4.FNumber", "nameField": "FFlex4.FName"},
            {"category": "dimensionDepartment", "categoryLabel": "核算维度·部门", "codeField": "FFlex5.FNumber", "nameField": "FFlex5.FName"},
            {"category": "dimensionCustomer", "categoryLabel": "核算维度·客户", "codeField": "FFlex6.FNumber", "nameField": "FFlex6.FName"},
            {"category": "dimensionEmployee", "categoryLabel": "核算维度·员工", "codeField": "FFlex7.FNumber", "nameField": "FFlex7.FName"},
            {"category": "dimensionMaterial", "categoryLabel": "核算维度·物料", "codeField": "FFlex8.FNumber", "nameField": "FFlex8.FName"},
            {"category": "dimensionExpense", "categoryLabel": "核算维度·费用项目", "codeField": "FFLEX9.FNumber", "nameField": "FFLEX9.FName"},
            {"category": "dimensionOrganization", "categoryLabel": "核算维度·组织机构", "codeField": "FFLEX11.FNumber", "nameField": "FFLEX11.FName"},
            {"category": "dimensionBank", "categoryLabel": "核算维度·银行", "codeField": "FFLEX14.FNumber", "nameField": "FFLEX14.FName"},
            {"category": "dimensionBankAccount", "categoryLabel": "核算维度·银行账号", "codeField": "FFLEX15.FNumber", "nameField": "FFLEX15.FName"},
            {"category": "dimensionOtherCounterparty", "categoryLabel": "核算维度·其他往来", "codeField": "FFLEX16.FNumber", "nameField": "FFLEX16.FName"},
            {"category": "dimensionServiceType", "categoryLabel": "核算维度·服务类型", "codeField": "FF100002.FNumber", "nameField": "FF100002.FName"},
            {"category": "dimensionUnit", "categoryLabel": "核算维度·Unit", "codeField": "FF100003", "nameField": "FF100003"},
            {"category": "dimensionRegion", "categoryLabel": "核算维度·入账地区", "codeField": "FF100004", "nameField": "FF100004"},
            {"category": "dimensionOldProject", "categoryLabel": "核算维度·旧项目", "codeField": "FF100006", "nameField": "FF100006"},
            {"category": "dimensionNewProject", "categoryLabel": "核算维度·新项目", "codeField": "FF100007", "nameField": "FF100007"},
        ],
    },
    {"category": "expense", "categoryLabel": "费用项目", "formId": "BD_Expense", "fields": ["FNumber", "FName"]},
    {"category": "currency", "categoryLabel": "币种", "formId": "BD_Currency", "fields": ["FNumber", "FName"]},
    {
        "category": "exchangeRate",
        "categoryLabel": "汇率",
        "formId": "BD_Rate",
        "fields": [
            "FRateID",
            "FRATETYPEID.FNumber",
            "FRATETYPEID.FName",
            "FBegDate",
            "FEndDate",
            "FCyForID.FNumber",
            "FCyForID.FName",
            "FCyToID.FNumber",
            "FCyToID.FName",
            "FExchangeRate",
            "FReverseExRate",
            "FDocumentStatus",
            "FForbidStatus",
        ],
        "codeField": "FRateID",
        "nameField": "FRATETYPEID.FName",
    },
    {"category": "taxRate", "categoryLabel": "税率", "formId": "BD_TaxRate", "fields": ["FNumber", "FName"]},
    {"category": "unit", "categoryLabel": "计量单位", "formId": "BD_UNIT", "fields": ["FNumber", "FName"]},
    {"category": "bank", "categoryLabel": "银行", "formId": "CN_BANK", "fields": ["FNumber", "FName"], "optional": True},
    {"category": "material", "categoryLabel": "物料", "formId": "BD_MATERIAL", "fields": ["FNumber", "FName"]},
    {"category": "stock", "categoryLabel": "仓库", "formId": "BD_STOCK", "fields": ["FNumber", "FName"]},
]

CAPABILITY_CATALOG: dict[str, Any] = {
    "version": "2026.07.1",
    "targets": [
        {
            "id": "kingdee-k3cloud",
            "brand": "金蝶",
            "product": "云星空",
            "versions": ["WebAPI v6+"],
            "adapter": "kingdee-k3cloud-webapi-v6",
            "modes": ["api", "template"],
            "verification": "contract-tested",
            "capabilities": [
                "probe",
                "sync_master_data",
                "check_period",
                "save_voucher_draft",
                "query_voucher",
                "query_by_idempotency_reference",
            ],
        },
        {
            "id": "yonyou-u8",
            "brand": "用友",
            "product": "U8",
            "versions": ["V12+"],
            "adapter": "yonyou-u8-openapi-v12",
            "modes": ["api", "template"],
            "verification": "configuration-required",
            "capabilities": [
                "probe",
                "sync_master_data",
                "check_period",
                "save_voucher_draft",
                "query_voucher",
                "query_by_idempotency_reference",
            ],
        },
        {
            "id": "inspur-gscloud",
            "brand": "浪潮",
            "product": "海岳 GS Cloud",
            "versions": ["iGIX/OpenAPI"],
            "adapter": "inspur-gscloud-igix",
            "modes": ["api", "template"],
            "verification": "configuration-required",
            "capabilities": [
                "probe",
                "sync_master_data",
                "check_period",
                "save_voucher_draft",
                "query_voucher",
                "query_by_idempotency_reference",
            ],
        },
        {
            "id": "other-file-target",
            "brand": "其他品牌",
            "product": "模板导入",
            "versions": ["客户提供"],
            "adapter": "file-template",
            "modes": ["template"],
            "verification": "customer-evidence-required",
            "capabilities": ["template_preview", "template_validate", "export_file"],
        },
    ],
    "sources": [
        {
            "id": "feishu-oa-json",
            "brand": "飞书 / Lark",
            "product": "审批",
            "versions": ["审批 v4"],
            "adapter": "feishu-approval-v4",
            "modes": ["api"],
            "verification": "configuration-required",
            "capabilities": [
                "probe",
                "approval_definition_read",
                "approval_incremental_sync",
                "field_mapping",
            ],
        },
        {
            "id": "wecom-oa-json",
            "brand": "企业微信",
            "product": "审批 API（JSON）",
            "versions": ["企业微信 API"],
            "adapter": "oa-json-api",
            "modes": ["api"],
            "verification": "configuration-required",
            "capabilities": ["probe", "json_incremental_sync", "field_mapping"],
        },
        {
            "id": "dingtalk-oa-json",
            "brand": "钉钉",
            "product": "审批 API（JSON）",
            "versions": ["开放平台"],
            "adapter": "oa-json-api",
            "modes": ["api"],
            "verification": "configuration-required",
            "capabilities": ["probe", "json_incremental_sync", "field_mapping"],
        },
        {
            "id": "other-oa-json",
            "brand": "其他 OA",
            "product": "通用 API（JSON）",
            "versions": ["客户接口"],
            "adapter": "oa-json-api",
            "modes": ["api"],
            "verification": "configuration-required",
            "capabilities": ["probe", "json_incremental_sync", "field_mapping"],
        },
        {
            "id": "oa-api-json",
            "brand": "OA 系统",
            "product": "API（JSON，旧版配置）",
            "versions": ["通用"],
            "adapter": "oa-json-api",
            "modes": ["api"],
            "verification": "configuration-required",
            "capabilities": ["probe", "json_incremental_sync", "field_mapping"],
            "hidden": True,
        },
        {
            "id": "feishu-approval",
            "brand": "飞书",
            "product": "审批（旧版专用连接）",
            "versions": ["v4"],
            "adapter": "feishu-approval-v4",
            "modes": ["api"],
            "verification": "contract-tested",
            "capabilities": ["probe", "approval_incremental_sync"],
            "hidden": True,
        },
        {
            "id": "local-files",
            "brand": "文件",
            "product": "本地文件",
            "versions": ["CSV", "XLSX", "XML", "XBRL", "OFD", "PDF", "图片", "银行流水"],
            "adapter": "local-files",
            "modes": ["file"],
            "verification": "built-in",
            "capabilities": ["archive", "parse", "ocr_candidates", "deduplicate"],
        },
    ],
}


def empty_production_state() -> dict[str, Any]:
    return {
        "version": STATE_VERSION,
        "operator": "",
        "operatorConfigured": False,
        "company": "",
        "ledger": "",
        "environment": "",
        "enterpriseProfiles": [],
        "targetSystem": None,
        "sourceSystems": [],
        "flowPlan": None,
        "readiness": {
            "plan": {"status": "not_ready", "validatedAt": None, "reasons": ["尚未生成接入方案"]},
            "systems": {"status": "not_ready", "validatedAt": None, "reasons": ["尚未完成系统与数据验证"]},
            "rules": {"status": "not_ready", "validatedAt": None, "reasons": ["尚未确认凭证场景"]},
            "production": {"status": "not_ready", "validatedAt": None, "reasons": ["尚未完成生产启用验证"]},
        },
        "productionActivation": {
            "enabled": False,
            "activatedAt": None,
            "confirmedTarget": None,
            "batchLimit": None,
        },
        "launchEvidence": {
            "permissionChecked": False,
            "periodChecked": False,
            "testDraftSaved": False,
            "voucherRechecked": False,
            "idempotencyChecked": False,
            "shadowRunChecked": False,
            "financeConfirmed": False,
        },
        "templateProfiles": [],
        "sourceDocuments": [],
        "events": [],
        "vouchers": [],
        "exceptions": [],
        "rules": [],
        "connectors": [],
        "masterData": default_account_master_data(),
        "defaultAccountsInitialized": True,
        "defaultAccountSource": dict(DEFAULT_ACCOUNT_SOURCE),
        "mappingTemplates": [],
        "approvalProcessingRules": [],
        "approvalUnionSelections": {},
        "approvalProcessingConfirmations": {},
        "auditLog": [],
        "syncLog": [],
        "externalQueryCache": [],
        "externalReadCache": [],
        "postingAttempts": [],
        "outbox": [],
        "activeFinanceConnectorId": "",
        "activeWorkflowConnectorId": "",
        "selectedEventId": "",
    }


def ensure_state_v2(state: dict[str, Any]) -> bool:
    changed = False
    blank = empty_production_state()
    for key, value in blank.items():
        if key not in state:
            state[key] = json.loads(json.dumps(value, ensure_ascii=False))
            changed = True
    if state.get("version") != STATE_VERSION:
        state["version"] = STATE_VERSION
        changed = True
    if initialize_default_accounts(state):
        changed = True
    for connector in state.get("connectors", []):
        if (
            connector.get("id") == "feishu-oa-json"
            and connector.get("name") != "飞书 / Lark 审批"
        ):
            connector["name"] = "飞书 / Lark 审批"
            changed = True
        if (
            connector.get("id") == "feishu-oa-json"
            and connector.get("adapter") == "oa-json-api"
        ):
            connector["adapter"] = "feishu-approval-v4"
            connector["status"] = "not_configured"
            connector["capabilities"] = []
            connector["lastProbe"] = None
            changed = True
        if connector.get("adapter") == "feishu-approval-v4":
            platform = str(connector.get("platform") or "").strip().lower()
            if platform not in {"feishu", "lark"}:
                platform = (
                    "lark"
                    if "larksuite.com" in str(connector.get("baseUrl") or "").lower()
                    else "feishu"
                )
                connector["platform"] = platform
                changed = True
            expected_base_url = (
                "https://open.larksuite.com"
                if platform == "lark"
                else "https://open.feishu.cn"
            )
            if connector.get("baseUrl") != expected_base_url:
                connector["baseUrl"] = expected_base_url
                changed = True
            for key, default in (
                ("appId", ""),
                ("approvalCode", ""),
                ("queryDateFrom", ""),
                ("queryDateTo", ""),
                ("approvalFields", []),
                ("fieldMapping", {}),
                ("fieldSources", []),
                ("additionalApprovalFieldIds", []),
                ("syncCursor", {}),
            ):
                if key not in connector:
                    connector[key] = json.loads(json.dumps(default, ensure_ascii=False))
                    changed = True
            if "approvalProfiles" not in connector:
                legacy_profile = legacy_approval_profile(connector)
                connector["approvalProfiles"] = [legacy_profile] if legacy_profile else []
                changed = True
            for profile in connector.get("approvalProfiles", []):
                if not isinstance(profile, dict):
                    continue
                approval_code = str(profile.get("approvalCode") or "").strip()
                if not profile.get("id") and approval_code:
                    profile["id"] = approval_profile_id(approval_code)
                    changed = True
                for key, default in APPROVAL_PROFILE_DEFAULTS.items():
                    if key not in profile:
                        profile[key] = json.loads(json.dumps(default, ensure_ascii=False))
                        changed = True
        if connector.get("adapter") != "kingdee-k3cloud-webapi-v6":
            continue
        if "serverUrl" not in connector:
            connector["serverUrl"] = connector.pop("baseUrl", "")
            changed = True
        if "acctId" not in connector:
            connector["acctId"] = connector.pop("accountId", "")
            changed = True
        if connector.get("authMode") != "app-id-secret-v3":
            connector["authMode"] = "app-id-secret-v3"
            connector["appId"] = str(connector.get("appId") or "")
            connector["orgNum"] = str(connector.get("orgNum") or "80016")
            connector["status"] = "not_configured"
            connector["capabilities"] = []
            connector["lastProbe"] = None
            changed = True
        queries = connector.setdefault("masterDataQueries", [])
        for query in queries:
            if query.get("formId") == "FIN_OTHERS" and query.get("category") == "counterparty":
                query["category"] = "otherCounterparty"
                query["categoryLabel"] = "其他往来"
                changed = True
            if query.get("formId") == "BOS_ASSISTANTDATA_DETAIL":
                assistant_default = next(
                    item for item in KINGDEE_MASTER_DATA_QUERIES
                    if item["formId"] == "BOS_ASSISTANTDATA_DETAIL"
                )
                for key in ("idFields", "codeField", "nameField", "replaceLegacyIdentity", "categoryLabel"):
                    if query.get(key) != assistant_default[key]:
                        query[key] = json.loads(json.dumps(assistant_default[key], ensure_ascii=False))
                        changed = True
                if "idField" in query:
                    query.pop("idField")
                    changed = True
            if query.get("formId") == "BD_FLEXITEMDETAILV":
                dimension_default = next(
                    item for item in KINGDEE_MASTER_DATA_QUERIES
                    if item["formId"] == "BD_FLEXITEMDETAILV"
                )
                for key in ("fields", "dimensionMappings", "categoryLabel"):
                    if query.get(key) != dimension_default[key]:
                        query[key] = json.loads(json.dumps(dimension_default[key], ensure_ascii=False))
                        changed = True
            if query.get("formId") in {"BD_Project", "CN_BANK"} and query.get("optional") is not True:
                query["optional"] = True
                changed = True
        known_form_ids = {str(query.get("formId") or "") for query in queries}
        for default_query in KINGDEE_MASTER_DATA_QUERIES:
            if default_query["formId"] in known_form_ids:
                continue
            queries.append(json.loads(json.dumps(default_query, ensure_ascii=False)))
            known_form_ids.add(default_query["formId"])
            changed = True
    return changed


def connector_template(system_id: str, *, environment: str = "测试环境") -> dict[str, Any]:
    entry = next(
        (item for item in CAPABILITY_CATALOG["targets"] + CAPABILITY_CATALOG["sources"] if item["id"] == system_id),
        None,
    )
    if not entry or entry["adapter"] in {"file-template", "local-files"}:
        raise ValueError("该系统不需要 API 连接器")
    source_ids = {item["id"] for item in CAPABILITY_CATALOG["sources"]}
    common = {
        "id": entry["id"],
        "name": f"{entry['brand']} {entry['product']}",
        "type": "workflow" if system_id in source_ids else "finance",
        "adapter": entry["adapter"],
        "environment": environment,
        "status": "not_configured",
        "baseUrl": "https://open.feishu.cn" if entry["adapter"] == "feishu-approval-v4" else "",
        "capabilities": [],
        "leastPrivilegeConfirmed": False,
    }
    if entry["adapter"] == "feishu-approval-v4":
        common.update({
            "platform": "feishu",
            "appId": "",
            "approvalCode": "",
            "approvalName": "",
            "queryDateFrom": "",
            "queryDateTo": "",
            "approvalFields": [],
            "fieldMapping": {},
            "fieldSources": [],
            "additionalApprovalFieldIds": [],
            "syncCursor": {},
            "approvalProfiles": [],
        })
    elif entry["adapter"] == "oa-json-api":
        common.update({
            "providerName": entry["brand"] if system_id != "other-oa-json" else "",
            "recordsPath": "data.items",
            "externalIdPath": "id",
            "approvalStatusPath": "status",
            "approvedValues": ["APPROVED", "approved", "已通过"],
            "authHeader": "Authorization",
            "authScheme": "Bearer",
            "fieldMapping": {},
            "syncCursor": {},
        })
    elif system_id == "kingdee-k3cloud":
        common.pop("baseUrl", None)
        common.update({
            "authMode": "app-id-secret-v3",
            "serverUrl": "",
            "acctId": "",
            "username": "",
            "appId": "",
            "orgNum": "80016",
            "ledger": "",
            "voucherFormId": "GL_VOUCHER",
            "voucherGroup": "PZZ47",
            "currencyCode": "PRE001",
            "exchangeRateType": "001",
            "localeId": 2052,
            "connectTimeout": 120,
            "requestTimeout": 120,
            "approvalControlEnabled": True,
            "enforceTargetMasterData": True,
            "enforcePeriodQuery": True,
            "periodQuery": {},
            "dimensionFieldMap": {},
            "masterDataQueries": json.loads(json.dumps(KINGDEE_MASTER_DATA_QUERIES, ensure_ascii=False)),
        })
    else:
        common.update({
            "accountId": "",
            "username": "",
            "ledger": "",
            "approvalControlEnabled": True,
            "enforceTargetMasterData": True,
            "enforcePeriodQuery": True,
            "endpointProfile": {},
            "fieldProfile": {},
            "dimensionFieldMap": {},
            "masterDataQueries": [],
        })
    return common


def invalidate_downstream(state: dict[str, Any], from_gate: str) -> None:
    order = ["plan", "systems", "rules", "production"]
    start = order.index(from_gate)
    for gate in order[start:]:
        state["readiness"][gate] = {
            "status": "not_ready",
            "validatedAt": None,
            "reasons": ["上游配置已变化，需要重新验证"],
        }
    state["productionActivation"] = {
        "enabled": False,
        "activatedAt": None,
        "confirmedTarget": None,
        "batchLimit": None,
    }


def invalidate_if_upstream_changed(existing: dict[str, Any] | None, incoming: dict[str, Any]) -> None:
    if not existing:
        return
    sections = [
        ("plan", ("enterpriseProfiles", "targetSystem", "sourceSystems")),
        ("systems", ("connectors", "masterData", "templateProfiles")),
        ("rules", ("rules",)),
    ]
    for gate, keys in sections:
        before = json.dumps({key: existing.get(key) for key in keys}, ensure_ascii=False, sort_keys=True)
        after = json.dumps({key: incoming.get(key) for key in keys}, ensure_ascii=False, sort_keys=True)
        if hashlib.sha256(before.encode()).digest() != hashlib.sha256(after.encode()).digest():
            invalidate_downstream(incoming, gate)
            return


class SetupService:
    def __init__(self, database: Database) -> None:
        self.database = database

    def _state(self) -> dict[str, Any]:
        state = self.database.get_state()
        if state is None:
            state = empty_production_state()
            self.database.put_state(state)
        ensure_state_v2(state)
        return state

    def plan(self, payload: dict[str, Any]) -> dict[str, Any]:
        target_id = str(payload.get("targetSystemId") or "").strip()
        source_ids = [str(item) for item in payload.get("sourceSystemIds") or []]
        target = next((item for item in CAPABILITY_CATALOG["targets"] if item["id"] == target_id), None)
        if not target:
            raise ValueError("请选择能力目录中的目标系统")
        sources = [item for item in CAPABILITY_CATALOG["sources"] if item["id"] in source_ids]
        if not sources:
            raise ValueError("至少选择一个数据来源")

        mode = "api" if "api" in target["modes"] else "template"
        blockers = []
        if target["verification"] in {"configuration-required", "customer-evidence-required"}:
            blockers.append("待客户提供目标版本接口或成功模板资料后验证")
        flow_steps = [
            {"order": 1, "name": "数据来源", "detail": "、".join(item["product"] for item in sources)},
            {"order": 2, "name": "资料匹配", "detail": "关联主数据、业务单据和审批依据"},
            {"order": 3, "name": "异常检查", "detail": "完整性、审批、期间、科目和辅助核算检查"},
            {"order": 4, "name": "凭证草稿", "detail": "仅按已启用的确定性规则生成"},
            {"order": 5, "name": "人工复核", "detail": "财务确认分录与来源证据"},
            {"order": 6, "name": "ERP 草稿/模板", "detail": "API 保存草稿" if mode == "api" else "生成已验证模板文件"},
            {"order": 7, "name": "回查", "detail": "取得外部编号并核对幂等引用" if mode == "api" else "记录测试账套导入结果"},
        ]
        state = self._state()
        state["targetSystem"] = {
            **target,
            "connectionMode": mode,
        }
        state["sourceSystems"] = sources
        state.pop("businessScenarios", None)
        state["flowPlan"] = {
            "catalogVersion": CAPABILITY_CATALOG["version"],
            "generatedAt": utc_now(),
            "mode": mode,
            "steps": flow_steps,
            "blockers": blockers,
            "manualNodes": ["异常处理", "凭证复核", "生产启用确认"],
        }
        invalidate_downstream(state, "plan")
        state["readiness"]["plan"] = {
            "status": "ready",
            "validatedAt": utc_now(),
            "reasons": blockers,
        }
        wanted = [target_id, *source_ids]
        state["connectors"] = [
            connector for connector in state["connectors"] if connector.get("id") in wanted
        ]
        existing_ids = {item.get("id") for item in state["connectors"]}
        for system_id in wanted:
            if system_id not in existing_ids and system_id not in {"local-files", "other-file-target"}:
                state["connectors"].append(connector_template(system_id))
        state["activeFinanceConnectorId"] = target_id if mode == "api" else ""
        workflow_connector = next(
            (item for item in state["connectors"] if item.get("type") == "workflow"),
            None,
        )
        state["activeWorkflowConnectorId"] = workflow_connector.get("id", "") if workflow_connector else ""
        self.database.put_state(state)
        return {"state": state, "plan": state["flowPlan"]}

    def preflight(self) -> dict[str, Any]:
        state = self._state()
        target = state.get("targetSystem") or {}
        mode = target.get("connectionMode")
        target_connector = next(
            (item for item in state["connectors"] if item.get("id") == target.get("id")),
            None,
        )
        template = next(
            (item for item in state["templateProfiles"] if item.get("targetSystemId") == target.get("id")),
            None,
        )
        system_reasons = []
        if mode == "api":
            if not target_connector or target_connector.get("status") != "connected":
                system_reasons.append("目标 ERP API 尚未测试通过")
            elif not (target_connector.get("lastProbe") or {}).get("ok"):
                system_reasons.append("目标 ERP 连接测试已失效")
        elif not template or template.get("testImportStatus") != "passed":
            system_reasons.append("目标 ERP 模板尚未在测试账套导入成功")
        for source in state.get("sourceSystems", []):
            if source.get("modes") == ["api"]:
                connector = next((item for item in state["connectors"] if item.get("id") == source["id"]), None)
                if not connector or connector.get("status") != "connected":
                    system_reasons.append(f"{source.get('brand')} {source.get('product')} 连接尚未测试通过")
        if not state.get("masterData"):
            system_reasons.append("尚未导入或同步目标基础资料")
        evidence_labels = {
            "permissionChecked": "尚未确认最小权限测试",
            "periodChecked": "尚未确认测试期间检查",
            "testDraftSaved": "尚未在测试账套保存凭证草稿",
            "voucherRechecked": "尚未回查取得真实外部编号",
            "idempotencyChecked": "尚未完成幂等重复请求测试",
            "shadowRunChecked": "尚未完成影子运行",
            "financeConfirmed": "尚未取得财务负责人确认",
        }
        evidence = state.get("launchEvidence") or {}
        system_reasons.extend(message for key, message in evidence_labels.items() if evidence.get(key) is not True)

        enabled_rules = [item for item in state.get("rules", []) if item.get("enabled")]
        rule_reasons = []
        if not enabled_rules:
            rule_reasons.append("尚无经人工确认并启用的凭证场景")
        for rule in enabled_rules:
            posting = rule.get("posting") or {}
            if not posting.get("debitAccountCode") or not posting.get("creditAccountCode"):
                rule_reasons.append(f"规则 {rule.get('name') or rule.get('id')} 缺少完整借贷科目")

        gates = {
            "plan": state["readiness"]["plan"],
            "systems": {
                "status": "ready" if not system_reasons else "not_ready",
                "validatedAt": utc_now(),
                "reasons": system_reasons,
            },
            "rules": {
                "status": "ready" if not rule_reasons else "not_ready",
                "validatedAt": utc_now(),
                "reasons": rule_reasons,
            },
            "production": {
                "status": "not_ready",
                "validatedAt": utc_now(),
                "reasons": ["需要明确确认目标公司、账套、环境和批量上限"],
            },
        }
        state["readiness"].update(gates)
        state["productionActivation"]["enabled"] = False
        self.database.put_state(state)
        return {"state": state, "ok": all(gates[key]["status"] == "ready" for key in ("plan", "systems", "rules")), "gates": gates}

    def activate(self, payload: dict[str, Any]) -> dict[str, Any]:
        report = self.preflight()
        if not report["ok"]:
            failed = next(
                reason
                for key in ("plan", "systems", "rules")
                for reason in report["gates"][key]["reasons"]
                if report["gates"][key]["status"] != "ready"
            )
            raise ValueError(f"生产启用前检查失败：{failed}")
        state = report["state"]
        enterprise = (state.get("enterpriseProfiles") or [{}])[0]
        expected = {
            "company": enterprise.get("legalEntity"),
            "accountSet": enterprise.get("accountSet"),
            "ledger": enterprise.get("ledger"),
            "environment": "生产环境",
        }
        for key, value in expected.items():
            if str(payload.get(key) or "").strip() != str(value or "").strip():
                raise ValueError(f"生产确认不一致：{key}")
        batch_limit = int(payload.get("batchLimit") or 0)
        if batch_limit < 1 or batch_limit > 1000:
            raise ValueError("生产批量上限必须在 1 到 1000 之间")
        state["environment"] = "生产环境"
        state["productionActivation"] = {
            "enabled": True,
            "activatedAt": utc_now(),
            "confirmedTarget": expected,
            "batchLimit": batch_limit,
            "environmentValidation": payload.get("environmentValidation"),
        }
        state["readiness"]["production"] = {
            "status": "ready",
            "validatedAt": utc_now(),
            "reasons": [],
        }
        self.database.put_state(state)
        return {"state": state, "activation": state["productionActivation"]}

    def preview_template(self, filename: str, content: bytes, target_system_id: str) -> dict[str, Any]:
        try:
            rows = parse_rows(filename, content)
        except (csv.Error, StopIteration):
            rows = []
        headers = headers_for(rows)
        suffix = Path(filename).suffix.lower()
        if not headers and suffix == ".csv":
            text = content.decode("utf-8-sig", errors="replace")
            headers = [str(item).strip() for item in next(csv.reader(io.StringIO(text)), []) if str(item).strip()]
        if not headers and suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("当前运行环境未安装 XLSX 解析组件 openpyxl") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            first = next(workbook.active.iter_rows(values_only=True), ())
            headers = [str(item).strip() for item in first if item not in (None, "")]
        if not headers:
            raise ValueError("模板没有可读取的表头")
        return {
            "filename": Path(filename).name,
            "targetSystemId": target_system_id,
            "headers": headers,
            "headerFingerprint": header_fingerprint(headers),
            "sampleRows": rows[:5],
            "rowCount": len(rows),
        }

    def validate_template(self, payload: dict[str, Any]) -> dict[str, Any]:
        state = self._state()
        headers = [str(item).strip() for item in payload.get("headers") or [] if str(item).strip()]
        required = [str(item).strip() for item in payload.get("requiredColumns") or [] if str(item).strip()]
        missing = [item for item in required if item not in headers]
        profile = {
            "id": str(payload.get("id") or f"TPL-{uuid.uuid4().hex[:10].upper()}"),
            "name": str(payload.get("name") or "未命名 ERP 模板").strip(),
            "targetSystemId": str(payload.get("targetSystemId") or "").strip(),
            "version": str(payload.get("version") or "1").strip(),
            "headerFingerprint": str(payload.get("headerFingerprint") or header_fingerprint(headers)),
            "headers": headers,
            "fieldMapping": payload.get("fieldMapping") or {},
            "requiredColumns": required,
            "formatRules": payload.get("formatRules") or {},
            "testImportStatus": str(payload.get("testImportStatus") or "not_tested"),
            "validatedAt": utc_now(),
            "validationErrors": [f"缺少必填列：{item}" for item in missing],
        }
        if not profile["targetSystemId"]:
            raise ValueError("模板档案缺少目标系统")
        existing = next((item for item in state["templateProfiles"] if item.get("id") == profile["id"]), None)
        if existing:
            existing.update(profile)
        else:
            state["templateProfiles"].append(profile)
        invalidate_downstream(state, "systems")
        self.database.put_state(state)
        return {"state": state, "profile": profile, "ok": not missing, "errors": profile["validationErrors"]}
