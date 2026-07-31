from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any


ALIASES = {
    "date": ("业务日期", "日期", "date", "business_date", "开票日期"),
    "counterparty": (
        "供应商",
        "客户",
        "客商",
        "交易对手主体",
        "counterparty",
        "vendor",
        "销售方名称",
        "购买方名称",
    ),
    "amount": ("付款金额", "金额", "含税金额", "amount", "gross_amount", "价税合计"),
    "reference": (
        "单据号",
        "审批单号",
        "Lark 编号",
        "订单号",
        "reference",
        "external_id",
        "发票号码",
    ),
    "company": ("公司", "主体", "company", "legal_entity", "购买方名称"),
    "ledger": ("账簿", "账套", "ledger", "book"),
    "business_type": ("业务类型", "business_type", "type"),
    "document_type": ("资料类型", "单据类型", "document_type", "source_type"),
    "invoice_no": ("发票号码", "发票号", "invoice_no", "invoice_number"),
    "seller_tax_id": ("销售方税号", "销方税号", "seller_tax_id"),
    "order_no": ("订单号", "采购订单号", "order_no", "po_number"),
    "approval_no": ("审批单号", "Lark 编号", "approval_no", "approval_number"),
    "bank_serial": ("流水号", "银行流水号", "唯一标识码", "bank_serial", "transaction_id"),
    "net_amount": ("不含税金额", "net_amount"),
    "tax_amount": ("税额", "tax_amount"),
    "payment_amount": ("付款金额", "实付金额", "payment_amount"),
    "currency": ("币别", "币种", "currency", "currency_code"),
    "exchange_rate": ("汇率", "记账汇率", "exchange_rate", "rate"),
    "department": ("部门", "department"),
    "project": ("项目", "project"),
    "summary": ("摘要", "业务摘要", "摘要（备注）", "summary", "货物或应税劳务名称"),
}

BANK_DEBIT_FIELDS = ("借方", "借方金额", "debit", "debit_amount")
BANK_CREDIT_FIELDS = ("贷方", "贷方金额", "credit", "credit_amount")
BANK_COUNTERPARTY_ACCOUNT_FIELDS = ("交易对手账号", "对方账号", "counterparty_account")
BANK_OWN_ACCOUNT_FIELDS = ("主体账号", "我方账号", "own_account")

STANDARD_FIELDS = {
    "date": "业务日期",
    "counterparty": "供应商",
    "amount": "含税金额",
    "reference": "审批单号",
    "company": "公司",
    "ledger": "账簿",
    "business_type": "业务类型",
    "document_type": "资料类型",
    "invoice_no": "发票号码",
    "seller_tax_id": "销售方税号",
    "order_no": "订单号",
    "approval_no": "审批单号",
    "bank_serial": "流水号",
    "net_amount": "不含税金额",
    "tax_amount": "税额",
    "payment_amount": "付款金额",
    "currency": "币别",
    "exchange_rate": "汇率",
    "department": "部门",
    "project": "项目",
    "summary": "摘要",
}


def parse_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        return parse_delimited_text(content)
    if suffix == ".xls":
        return parse_xls(content)
    if suffix == ".xlsx":
        return parse_xlsx(content)
    if suffix == ".xml":
        row = parse_xml(content)
        return [row] if row else []
    if suffix == ".xbrl":
        row = parse_xml(content)
        return [row] if row else []
    if suffix == ".ofd":
        row = parse_ofd(content)
        return [row] if row else []
    return []


def headers_for(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    return [str(header) for header in rows[0].keys()]


def header_fingerprint(headers: list[str]) -> str:
    normalized = "\n".join(sorted(header.strip().lower() for header in headers if header.strip()))
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized = {header.strip().lower(): header for header in headers}
    for target, aliases in ALIASES.items():
        for alias in aliases:
            if alias.lower() in normalized:
                mapping[target] = normalized[alias.lower()]
                break
    return mapping


def apply_mapping(row: dict[str, Any], mapping: dict[str, Any] | None) -> dict[str, Any]:
    if not mapping:
        return row
    mapped = dict(row)
    for target, source in mapping.items():
        if isinstance(source, list):
            if target != "amount":
                raise ValueError("只有付款金额支持选择多个源字段")
            sources = list(dict.fromkeys(str(item) for item in source if str(item).strip()))
            populated = [row[item] for item in sources if item in row and row[item] not in (None, "")]
            non_zero = []
            for value in populated:
                try:
                    is_zero = Decimal(str(value).replace(",", "").strip()) == 0
                except InvalidOperation:
                    is_zero = False
                if not is_zero:
                    non_zero.append(value)
            candidates = non_zero or populated
            if len(candidates) > 1:
                raise ValueError("付款金额所选字段在同一行存在多个值，请核对借方和贷方")
            if candidates:
                mapped[STANDARD_FIELDS["payment_amount"]] = candidates[0]
            continue
        canonical = STANDARD_FIELDS.get(target)
        if canonical and source in row:
            mapped[canonical] = row[source]
    return mapped


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    bank_serial = pick(row, ALIASES["bank_serial"])
    counterparty = pick(row, ALIASES["counterparty"])
    debit = pick(row, BANK_DEBIT_FIELDS)
    credit = pick(row, BANK_CREDIT_FIELDS)
    has_directional_columns = any(
        name in row for name in (*BANK_DEBIT_FIELDS, *BANK_CREDIT_FIELDS)
    )
    if not (bank_serial and counterparty and has_directional_columns):
        return normalized
    if debit not in (None, "") and credit not in (None, ""):
        raise ValueError("银行流水同一行不能同时存在借方和贷方金额")
    if debit in (None, "") and credit in (None, ""):
        raise ValueError("银行流水缺少借方或贷方金额")

    direction = "inflow" if debit not in (None, "") else "outflow"
    amount_cents = abs(to_cents(debit if direction == "inflow" else credit))
    normalized["付款金额"] = format(Decimal(amount_cents) / Decimal(100), "f")
    normalized["流水号"] = str(bank_serial).strip()
    normalized["资料类型"] = "银行流水"
    normalized["业务类型"] = "银行收款" if direction == "inflow" else "银行付款"
    normalized["收支方向"] = direction

    summary = pick(row, ALIASES["summary"])
    if summary not in (None, ""):
        normalized["摘要"] = str(summary).strip()
    counterparty_account = pick(row, BANK_COUNTERPARTY_ACCOUNT_FIELDS)
    if counterparty_account not in (None, ""):
        normalized["交易对手账号"] = str(counterparty_account).strip()
    own_account = pick(row, BANK_OWN_ACCOUNT_FIELDS)
    if own_account not in (None, ""):
        normalized["我方账号"] = str(own_account).strip()
    return normalized


def normalize_import_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows = [normalize_import_row(row) for row in rows]
    serial_counts: dict[str, int] = {}
    for row in normalized_rows:
        if row.get("资料类型") != "银行流水":
            continue
        serial = str(row.get("流水号") or "").strip()
        serial_counts[serial] = serial_counts.get(serial, 0) + 1

    digest_occurrences: dict[str, int] = {}
    for row in normalized_rows:
        if row.get("资料类型") != "银行流水":
            continue
        source_serial = str(row.get("流水号") or "").strip()
        if serial_counts.get(source_serial, 0) <= 1:
            continue
        row["源唯一标识码"] = source_serial
        fingerprint_row = dict(row)
        fingerprint_row.pop("流水号", None)
        fingerprint_row.pop("源唯一标识码", None)
        digest = hashlib.sha256(
            json.dumps(
                fingerprint_row,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()[:16].upper()
        occurrence = digest_occurrences.get(digest, 0) + 1
        digest_occurrences[digest] = occurrence
        row["流水号"] = f"BANK-{digest}" if occurrence == 1 else f"BANK-{digest}-{occurrence:02d}"
    return normalized_rows


def import_profile(rows: list[dict[str, Any]]) -> str:
    return "bankStatement" if any(
        row.get("资料类型") == "银行流水" for row in rows
    ) else "generic"


def parse_csv(content: bytes) -> list[dict[str, str]]:
    return parse_delimited_text(content)


def parse_delimited_text(content: bytes) -> list[dict[str, str]]:
    text = decode_text(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error as exc:
        raise ValueError("文本文件必须使用逗号、分号、制表符或竖线分隔列") from exc
    return [dict(row) for row in csv.DictReader(io.StringIO(text), dialect=dialect)]


def parse_xls(content: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("当前运行环境未安装 XLS 解析组件 xlrd") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        sheet = workbook.sheet_by_index(0)
    except (xlrd.XLRDError, IndexError) as exc:
        raise ValueError("XLS 文件损坏、加密或没有可读取的工作表") from exc
    if sheet.nrows == 0:
        return []

    def value_at(row: int, column: int) -> Any:
        if not hasattr(sheet, "cell"):
            return spreadsheet_value(sheet.cell_value(row, column))
        cell = sheet.cell(row, column)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return spreadsheet_value(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
        return spreadsheet_value(cell.value)

    headers = [str(value_at(0, column)).strip() for column in range(sheet.ncols)]
    result = []
    for row_index in range(1, sheet.nrows):
        values = [value_at(row_index, column) for column in range(sheet.ncols)]
        if not any(value not in (None, "") for value in values):
            continue
        result.append({header: value for header, value in zip(headers, values) if header})
    workbook.release_resources()
    return result


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("当前运行环境未安装 XLSX 解析组件 openpyxl") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [
        str(spreadsheet_value(value)).strip() if value is not None else ""
        for value in next(rows)
    ]
    return [
        {
            header: spreadsheet_value(value)
            for header, value in zip(headers, values)
            if header
        }
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def spreadsheet_value(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def parse_xml(content: bytes) -> dict[str, Any]:
    root = ET.fromstring(content)
    values: dict[str, str] = {}
    for element in root.iter():
        tag = re.sub(r"^\{.*\}", "", element.tag)
        text = (element.text or "").strip()
        if text and tag not in values:
            values[tag] = text
    normalized: dict[str, Any] = {}
    for target, names in ALIASES.items():
        for name in names:
            if name in values:
                normalized[name] = values[name]
                break
    return normalized


def parse_ofd(content: bytes) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        members = [
            info for info in archive.infolist()
            if not info.is_dir()
            and Path(info.filename).suffix.lower() in {".xml", ".xbrl"}
            and info.file_size <= 20 * 1024 * 1024
        ]
        if len(members) > 500:
            raise ValueError("OFD 内嵌结构化文件数量异常")
        if sum(info.file_size for info in members) > 100 * 1024 * 1024:
            raise ValueError("OFD 解压后的结构化数据超过 100 MB")
        for info in members:
            try:
                row = parse_xml(archive.read(info))
            except ET.ParseError:
                continue
            for key, value in row.items():
                merged.setdefault(key, value)
    return merged


def event_from_row(row: dict[str, Any], document_id: str, sequence: int, unique: str) -> dict[str, Any]:
    amount = pick(row, ALIASES["amount"])
    counterparty = pick(row, ALIASES["counterparty"])
    if amount in (None, "") or not str(counterparty).strip():
        raise ValueError("缺少“金额”或“供应商/客商”字段")
    amount_cents = to_cents(amount)
    business_date = normalize_date(pick(row, ALIASES["date"]))
    order_no = str(pick(row, ALIASES["order_no"]) or "").strip()
    approval_no = str(pick(row, ALIASES["approval_no"]) or "").strip()
    invoice_no = str(pick(row, ALIASES["invoice_no"]) or "").strip()
    bank_serial = str(pick(row, ALIASES["bank_serial"]) or "").strip()
    bank_direction = str(row.get("收支方向") or "").strip()
    bank_account = str(pick(row, BANK_COUNTERPARTY_ACCOUNT_FIELDS) or "").strip()
    own_bank_account = str(pick(row, BANK_OWN_ACCOUNT_FIELDS) or "").strip()
    reference = str(
        pick(row, ALIASES["reference"])
        or approval_no
        or order_no
        or invoice_no
        or bank_serial
        or f"IMP-{unique}-{sequence:04d}"
    )
    department = str(pick(row, ALIASES["department"]) or "")
    project = str(pick(row, ALIASES["project"]) or "")
    summary = str(pick(row, ALIASES["summary"]) or str(counterparty))
    company = str(pick(row, ALIASES["company"]) or "")
    ledger = str(pick(row, ALIASES["ledger"]) or "")
    business_type = str(pick(row, ALIASES["business_type"]) or "")
    document_type = str(pick(row, ALIASES["document_type"]) or "业务资料")
    currency = str(pick(row, ALIASES["currency"]) or "CNY").strip()
    exchange_rate_value = pick(row, ALIASES["exchange_rate"])
    exchange_rate = ""
    if exchange_rate_value not in (None, ""):
        try:
            parsed_rate = Decimal(str(exchange_rate_value).replace(",", "").strip())
        except InvalidOperation as exc:
            raise ValueError(f"汇率格式无效：{exchange_rate_value}") from exc
        if parsed_rate <= 0:
            raise ValueError("汇率必须大于零")
        exchange_rate = format(parsed_rate, "f")
    seller_tax_id = str(pick(row, ALIASES["seller_tax_id"]) or "").strip()
    business_identity = (
        f"bank|{bank_serial}"
        if bank_serial
        else f"{company}|{ledger}|{business_type}|{reference}"
    )
    business_key = hashlib.sha256(business_identity.encode("utf-8")).hexdigest()[:24]
    source_record_key = ""
    if invoice_no and seller_tax_id:
        source_record_key = f"invoice:{seller_tax_id}:{invoice_no}"
    elif bank_serial:
        source_record_key = f"bank:{bank_serial}"
    return {
        "id": f"EV-{unique}-{sequence}",
        "reference": reference,
        "businessKey": business_key,
        "type": business_type,
        "company": company,
        "ledger": ledger,
        "date": business_date,
        "counterparty": str(counterparty),
        "amountCents": amount_cents,
        "amountBreakdown": {
            "grossCents": amount_cents,
            "netCents": optional_cents(pick(row, ALIASES["net_amount"])),
            "taxCents": optional_cents(pick(row, ALIASES["tax_amount"])),
            "paymentCents": optional_cents(pick(row, ALIASES["payment_amount"])),
        },
        "currency": currency,
        "exchangeRate": exchange_rate,
        "approvalNo": approval_no,
        "bankSerial": bank_serial,
        "bankDirection": bank_direction,
        "bankAccount": bank_account,
        "ownBankAccount": own_bank_account,
        "department": department,
        "project": project,
        "summary": summary,
        "approvalStatus": "unknown",
        "sourceVerified": False,
        "financeReviewed": False,
        "pushAllowed": False,
        "matchConfidence": None,
        "sourceDocumentIds": [document_id],
        "sourceRecords": [{
            "documentId": document_id,
            "documentType": document_type,
            "recordKey": source_record_key,
            "referenceFields": {
                "reference": reference,
                "orderNo": order_no,
                "approvalNo": approval_no,
                "invoiceNo": invoice_no,
                "bankSerial": bank_serial,
                "bankDirection": bank_direction,
                "bankAccount": bank_account,
                "ownBankAccount": own_bank_account,
                "sellerTaxId": seller_tax_id,
            },
            "amountCents": amount_cents,
        }],
        "matchExplanation": [
            f"按业务唯一键 {business_type} / {reference} 关联",
            "交易对方和金额用于一致性校验",
        ],
        "exceptionIds": [],
        "status": "可生成",
    }


def pick(row: dict[str, Any], names: tuple[str, ...]) -> Any:
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return ""


def to_cents(value: Any) -> int:
    normalized = str(value).replace(",", "").replace("¥", "").replace("￥", "").strip()
    try:
        return int((Decimal(normalized) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except InvalidOperation as exc:
        raise ValueError(f"金额格式无效：{value}") from exc


def optional_cents(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return to_cents(value)


def normalize_date(value: Any) -> str:
    if not value:
        return date.today().isoformat()
    if hasattr(value, "date"):
        return value.date().isoformat()
    text = str(value).strip().replace("/", "-")
    numeric_text = text.lstrip("-")
    if numeric_text.replace(".", "", 1).isdigit():
        timestamp = float(text)
        if abs(timestamp) >= 100_000_000_000:
            timestamp /= 1000
        if abs(timestamp) >= 1_000_000_000:
            return datetime.fromtimestamp(timestamp, timezone.utc).date().isoformat()
    return text[:10]


def decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别 CSV/TXT 文本编码")
